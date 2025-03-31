import sys
import os
import time

import openpyxl
import pandas as pd
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
import json

from openpyxl import *
from openpyxl.styles import *


class RowColumnAdjustDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("调整行列间距")

        # 布局
        layout = QVBoxLayout(self)

        # 行范围输入框
        self.row_label = QLabel("请输入要调整的行范围（例如：1-5）:", self)
        self.row_input = QLineEdit(self)
        self.row_input.setText("1-5")  # 默认值为1-5

        # 行间距输入框
        self.row_spacing_label = QLabel("请输入新的行间距（行高）:", self)
        self.row_spacing_input = QLineEdit(self)
        self.row_spacing_input.setValidator(QIntValidator(1, 1000))  # 限制输入为正整数
        self.row_spacing_input.setText("20")  # 默认值为20

        # 列范围输入框
        self.col_label = QLabel("请输入要调整的列范围（例如：2-6）:", self)
        self.col_input = QLineEdit(self)
        self.col_input.setText("2-6")  # 默认值为2-6

        # 列间距输入框
        self.col_spacing_label = QLabel("请输入新的列间距（列宽）:", self)
        self.col_spacing_input = QLineEdit(self)
        self.col_spacing_input.setValidator(QIntValidator(1, 1000))  # 限制输入为正整数
        self.col_spacing_input.setText("100")  # 默认值为100

        # 确定按钮
        self.ok_button = QPushButton("确定", self)
        self.ok_button.clicked.connect(self.accept)

        # 布局添加控件
        layout.addWidget(self.row_label)
        layout.addWidget(self.row_input)
        layout.addWidget(self.row_spacing_label)
        layout.addWidget(self.row_spacing_input)
        layout.addWidget(self.col_label)
        layout.addWidget(self.col_input)
        layout.addWidget(self.col_spacing_label)
        layout.addWidget(self.col_spacing_input)
        layout.addWidget(self.ok_button)

    def get_values(self):
        try:
            # 获取行范围和列范围
            row_range = self.row_input.text().split('-')
            col_range = self.col_input.text().split('-')

            row_start = int(row_range[0]) - 2  # 转换为0索引
            row_end = int(row_range[1]) - 2  # 转换为0索引

            col_start = int(col_range[0]) - 1  # 转换为0索引
            col_end = int(col_range[1]) - 1  # 转换为0索引

            # 获取行间距和列间距
            row_height = int(self.row_spacing_input.text())
            col_width = int(self.col_spacing_input.text())

            return row_start, row_end, col_start, col_end, row_height, col_width
        except ValueError:
            return None, None, None, None, None, None  # 如果输入无效，返回None

class SimpleExcel(QMainWindow):
    def __init__(self):
        super().__init__()
        self.data_modified = False  # 用于追踪数据是否被修改
        self.setWindowTitle("简化版 Excel")
        self.setGeometry(100, 100, 800, 600)
        self.auto_save_enabled = False  # 初始时自动保存是禁用的
        self.data_modified = False  # 用于追踪数据是否被修改
        self.auto_save_timer = QTimer(self)  # 创建一个定时器
        self.auto_save_timer.timeout.connect(self.auto_save)  # 连接超时事件到自动保存方法
        # 初始化历史文件路径列表
        self.history_files = []
        # 创建主窗口部件
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        self.table = QTableWidget(100, 50)  # 100行50列        # 创建表格（初始为空）
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: #f4f6f9;
                gridline-color: #dfe3e8;
                border: none;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #3a8ee6;
                color: white;
            }
            QTableWidget::item:hover {
                background-color: #e6f1ff;
            }
            QHeaderView::section {
                background-color: #0078d4;
                color: white;
                font-weight: bold;
                padding: 5px;
                border: none;
            }
            QHeaderView::section:horizontal {
                border-right: 1px solid #dfe3e8;
            }
            QHeaderView::section:vertical {
                border-bottom: 1px solid #dfe3e8;
            }
            QTableWidget::horizontalHeader {
                height: 40px;
            }
            QTableWidget::verticalHeader {
                width: 50px;
            }
        """)

        # 设置字体

        # 设置表头
        # 创建表格（初始为空）

        self.table.setAlternatingRowColors(True)
        self.table.setHorizontalHeaderLabels([f"列 {i + 1}" for i in range(50)])
        self.table.itemChanged.connect(self.on_data_modified)
        self.table.itemSelectionChanged.connect(self.update_input_formula)  # 更新输入框内容

        # 布局
        layout = QVBoxLayout()
        layout.addWidget(self.table)
        main_widget.setLayout(layout)

        # 设置菜单
        self.create_menu()

        # 设置状态栏
        self.statusBar = self.statusBar()
        self.statusBar.showMessage("就绪")

        # 设置整体窗口的调色板
        palette = self.palette()
        palette.setColor(QPalette.Window, QColor("#ffffff"))
        palette.setColor(QPalette.WindowText, QColor("#343a40"))
        self.setPalette(palette)
        main_layout = QVBoxLayout(main_widget)
        # 创建图形视图层
        self.graphics_view = QGraphicsView()
        self.graphics_scene = QGraphicsScene(self)
        self.graphics_view.setScene(self.graphics_scene)
        main_layout.addWidget(self.graphics_view)
        self.input_formula = QLineEdit()
        self.input_formula.setPlaceholderText("在此输入公式或内容")
        self.input_formula.returnPressed.connect(self.apply_formula)  # 回车应用公式
        layout.addWidget(self.input_formula)
        layout.addWidget(self.table)
        main_widget.setLayout(layout)
        # 创建工具栏（开始菜单）
        self.create_toolbar()


    def create_toolbar(self):
        toolbar = self.addToolBar("开始菜单")
        toolbar.setMovable(False)  # 禁止移动工具栏

        # 设置工具栏的整体样式
        toolbar.setStyleSheet("""
            QToolBar {
                background-color: #ffffff;
                padding: 10px;
                border-radius: 10px;
                border: 1px solid #dcdcdc;
                margin: 10px;
            }
            QToolButton {
                background-color: #f0f0f0;
                color: #333333;
                padding: 5px 10px;
                margin: 5px;
                border-radius: 4px;
                border: 1px solid #cccccc;
            }
            QToolButton:hover {
                background-color: #e6e6e6;
                color: #000000;
            }
            QToolButton:pressed {
                background-color: #cccccc;
                color: #000000;
            }
        """)

        auto_save_action = QAction("自动保存", self, checkable=True)
        auto_save_action.setToolTip("启用/禁用自动保存")
        auto_save_action.triggered.connect(self.toggle_auto_save)  # 连接到开关函数
        toolbar.addAction(auto_save_action)

        # 添加字体加粗按钮
        bold_action = QAction("B", self)
        bold_action.setShortcut("Ctrl+B")
        bold_action.setToolTip("加粗")
        bold_action.triggered.connect(self.toggle_bold)
        toolbar.addAction(bold_action)

        # 添加斜体按钮
        italic_action = QAction("I", self)
        italic_action.setShortcut("Ctrl+I")
        italic_action.setToolTip("斜体")
        italic_action.triggered.connect(self.toggle_italic)
        toolbar.addAction(italic_action)

        # 添加居中对齐按钮
        center_action = QAction("C", self)
        center_action.setToolTip("居中对齐")
        center_action.triggered.connect(self.toggle_center_align)
        toolbar.addAction(center_action)

        # 添加字体选择按钮
        font_action = QAction("字体", self)
        font_action.triggered.connect(self.select_font)
        toolbar.addAction(font_action)

        # 添加字体大小按钮
        font_size_action = QAction("大小", self)
        font_size_action.triggered.connect(self.select_font_size)
        toolbar.addAction(font_size_action)

        # 添加增加行列按钮
        add_row_action = QAction("增加行", self)
        add_row_action.triggered.connect(self.new_rows)
        toolbar.addAction(add_row_action)

        add_col_action = QAction("增加列", self)
        add_col_action.triggered.connect(self.new_cols)
        toolbar.addAction(add_col_action)

        toolbar = self.addToolBar("插入")

        # 插入图片
        insert_image_action = QAction("插入图片", self)
        insert_image_action.triggered.connect(self.insert_image)
        toolbar.addAction(insert_image_action)

        # 插入文本框
        insert_textbox_action = QAction("插入文本框", self)
        insert_textbox_action.triggered.connect(self.insert_textbox)
        toolbar.addAction(insert_textbox_action)

        # 插入艺术字
        insert_wordart_action = QAction("插入艺术字", self)
        insert_wordart_action.triggered.connect(self.insert_wordart)
        toolbar.addAction(insert_wordart_action)

    # 更新输入框内容，显示选中单元格的内容或公式
    def update_input_formula(self):
        selected_item = self.table.currentItem()
        if selected_item:
            formula = selected_item.data(Qt.UserRole)
            if formula and formula.startswith("="):
                self.input_formula.setText(formula)  # 显示公式在顶部输入框
            else:
                self.input_formula.setText(selected_item.text())  # 非公式内容显示值
        else:
            self.input_formula.clear()

    # 将输入框内容应用到选中的单元格
    def apply_formula(self):
        selected_item = self.table.currentItem()
        if selected_item:
            formula = self.input_formula.text().strip()
            selected_item.setData(Qt.UserRole, formula)  # 存储公式到 UserRole
            if formula.startswith("="):  # 如果是公式
                try:
                    expression = formula[1:]
                    result = eval(expression, {"__builtins__": None}, {})
                    selected_item.setText(str(result))  # 单元格显示计算结果
                    self.input_formula.setText(formula)  # 输入框保持公式
                except Exception as e:
                    self.statusBar.showMessage(f"计算错误: {e}", 5000)
            else:
                selected_item.setText(formula)  # 非公式内容直接显示

    # 数据修改时触发
    def on_data_modified(self, item):
        text = item.text().strip()
        if text.startswith("="):  # 如果输入的是公式
            item.setData(Qt.UserRole, text)  # 存储公式到 UserRole
            try:
                expression = text[1:]
                result = eval(expression, {"__builtins__": None}, {})
                item.setText(str(result))  # 显示计算结果
            except Exception as e:
                self.statusBar.showMessage(f"计算错误: {e}", 5000)
        else:
            item.setData(Qt.UserRole, text)  # 非公式内容直接保存

    def insert_image(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择图片", "", "Images (*.png *.jpg *.bmp)")
        if file_path:
            pixmap = QPixmap(file_path)
            item = QGraphicsPixmapItem(pixmap)
            self.graphics_scene.addItem(item)

    def on_data_modified(self, item):
        self.data_modified = True
        # 获取单元格内容
        content = item.text().strip()

        # 检查内容是否以 "=" 开头
        if content.startswith("="):
            try:
                # 计算公式，去掉 "="
                expression = content[1:]
                # 使用安全的 eval 进行计算
                result = eval(expression, {"__builtins__": None}, {})
                item.setText(str(result))  # 更新单元格为计算结果
            except Exception as e:
                self.statusBar.showMessage(f"计算错误: {e}", 5000)
        else:
            # 标记为已修改
            self.data_modified = True

    def insert_textbox(self):
        text, ok = QInputDialog.getText(self, "插入文本框", "输入文本内容:")
        if ok and text:
            label = QLabel(text)
            label.setStyleSheet("background-color: #f0f0f0; border: 1px solid #cccccc; padding: 5px;")
            self.graphics_scene.addWidget(label)

    def insert_wordart(self):
        text, ok = QInputDialog.getText(self, "插入艺术字", "输入艺术字内容:")
        if ok and text:
            font, ok = QInputDialog.setFont(self, "选择艺术字字体")
            if ok:
                item = QGraphicsTextItem(text)
                item.setFont(font)
                item.setDefaultTextColor(QColor("#0078d4"))
                self.graphics_scene.addItem(item)
    def toggle_auto_save(self, checked):
        """切换自动保存的开关"""
        if checked:
            self.auto_save_enabled = True
            self.auto_save_timer.start(60000)  # 每60秒自动保存一次
            self.statusBar.showMessage(f"自动保存已开启！", 5000)
            if self.auto_save_enabled:
                timestamp = time.strftime("%Y%m%d_%H%M%S")  # 使用当前时间戳来避免文件名冲突
                new_path = f"auto_save_{timestamp}.xlsx"
                self.file_path = fr"C:\Users\Public\{new_path}"  # 可以设置为一个固定路径或当前文件路径
                file_path = self.file_path

                if os.path.exists(file_path):  # 判断文件是否存在
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        self.statusBar.showMessage(f"在自动保存时失败: {str(e)}")
                        QMessageBox.warning(self, "错误", f"在设置自动保存时失败: {str(e)}")

                if os.path.exists(file_path) == False:  # 判断文件是否存在
                    try:
                        # 这里添加您保存数据到文件的代码
                        # 例如，使用 pandas 将数据保存到文件
                        self.path_save_excel(file_path)
                    except Exception as e:
                        self.statusBar.showMessage(f"在设置自动保存时失败: {str(e)}")
                        QMessageBox.warning(self, "错误", f"在设置自动保存时失败: {str(e)}")
        else:
            self.auto_save_enabled = False
            self.auto_save_timer.stop()  # 停止定时器
            self.statusBar.showMessage(f"已关闭自动保存", 5000)

    def auto_save(self):
        """自动保存文件"""
        if self.auto_save_enabled:
            file_path = self.file_path
            if os.path.exists(file_path):  # 判断文件是否存在
                try:
                    os.remove(file_path)
                except Exception as e:
                    self.statusBar.showMessage(f"在自动保存时失败: {str(e)}")
                    QMessageBox.warning(self, "错误", f"自动保存时失败: {str(e)}")


            if os.path.exists(file_path) == False:  # 判断文件是否存在
                try:
                    # 这里添加您保存数据到文件的代码
                    print(f"自动保存成功！")
                    # 例如，使用 pandas 将数据保存到文件
                    self.path_save_excel(file_path)
                except Exception as e:
                    self.statusBar.showMessage(f"在自动保存时失败: {str(e)}")

    def apply_bold(self):
        # 应用粗体样式到选中的单元格
        self.apply_formatting(bold=self.bold_action.isChecked())

    def apply_italic(self):
        # 应用斜体样式到选中的单元格
        self.apply_formatting(italic=self.italic_action.isChecked())

    def apply_underline(self):
        # 应用下划线样式到选中的单元格
        self.apply_formatting(underline=self.underline_action.isChecked())

    def apply_formatting(self, bold=None, italic=None, underline=None):
        # 获取选中的单元格，并应用相应的样式
        for item in self.table.selectedItems():
            font = item.font()
            if bold is not None:
                font.setBold(bold)
            if italic is not None:
                font.setItalic(italic)
            if underline is not None:
                font.setUnderline(underline)
            item.setFont(font)

    def toggle_bold(self):
        selected_ranges = self.table.selectedRanges()
        if selected_ranges:
            for selection_range in selected_ranges:
                for row in range(selection_range.topRow(), selection_range.bottomRow() + 1):
                    for col in range(selection_range.leftColumn(), selection_range.rightColumn() + 1):
                        item = self.table.item(row, col)
                        if item:
                            current_font = item.font()
                            current_font.setBold(not current_font.bold())
                            item.setFont(current_font)


    def toggle_italic(self):
        selected_ranges = self.table.selectedRanges()
        if selected_ranges:
            for selection_range in selected_ranges:
                for row in range(selection_range.topRow(), selection_range.bottomRow() + 1):
                    for col in range(selection_range.leftColumn(), selection_range.rightColumn() + 1):
                        item = self.table.item(row, col)
                        if item:
                            current_font = item.font()
                            current_font.setItalic(not current_font.italic())
                            item.setFont(current_font)

    def toggle_center_align(self):
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            QMessageBox.warning(self, "警告", "没有选中任何单元格!")
            return

        for selection_range in selected_ranges:
            for row in range(selection_range.topRow(), selection_range.bottomRow() + 1):
                for col in range(selection_range.leftColumn(), selection_range.rightColumn() + 1):
                    item = self.table.item(row, col)
                    if item:
                        current_alignment = item.textAlignment()
                        if current_alignment == Qt.AlignCenter:
                            item.setTextAlignment(Qt.AlignLeft)
                        else:
                            item.setTextAlignment(Qt.AlignCenter)

    def select_font(self):
        font, ok = QFontDialog.getFont()
        if ok:
            selected_ranges = self.table.selectedRanges()
            for selection_range in selected_ranges:
                for row in range(selection_range.topRow(), selection_range.bottomRow() + 1):
                    for col in range(selection_range.leftColumn(), selection_range.rightColumn() + 1):
                        item = self.table.item(row, col)
                        if item:
                            item.setFont(font)

    def select_font_size(self):
        size, ok = QInputDialog.getInt(self, "选择字体大小", "请输入字体大小:", min=1, max=100, value=10)
        if ok:
            selected_ranges = self.table.selectedRanges()
            for selection_range in selected_ranges:
                for row in range(selection_range.topRow(), selection_range.bottomRow() + 1):
                    for col in range(selection_range.leftColumn(), selection_range.rightColumn() + 1):
                        item = self.table.item(row, col)
                        if item:
                            current_font = item.font()
                            current_font.setPointSize(size)
                            item.setFont(current_font)

    def new_rows(self):
        num_rows, ok = QInputDialog.getInt(self, "批量新建行", "请输入要新建的行数：", 1, 1, 100, 1)
        if ok:
            current_row_count = self.table.rowCount()
            self.table.setRowCount(current_row_count + num_rows)
            self.statusBar.showMessage(f"已新增 {num_rows} 行", 5000)

    def new_cols(self):
        num_cols, ok = QInputDialog.getInt(self, "批量新建列", "请输入要新建的列数：", 1, 1, 100, 1)
        if ok:
            current_col_count = self.table.columnCount()
            self.table.setColumnCount(current_col_count + num_cols)
            self.statusBar.showMessage(f"已新增 {num_cols} 列", 5000)

    def get_history_file_path(self):
        # 获取用户名的环境变量，并构造历史文件路径
        user_name = os.getenv("USERNAME")
        folder_path = os.path.join("C:\\Users", user_name, "AppData\\Local\\Easy Excel")
        os.makedirs(folder_path, exist_ok=True)  # 如果文件夹不存在则创建
        return os.path.join(folder_path, "history.json")

    def save_to_history(self, file_path):
        # 读取现有的历史记录，如果不存在则初始化为空列表
        history_file_path = self.get_history_file_path()
        if os.path.exists(history_file_path):
            with open(history_file_path, "r", encoding="utf-8") as file:
                history = json.load(file)
        else:
            history = []

        # 将新文件路径添加到历史记录并写入文件
        if file_path not in history:
            history.insert(0, file_path)  # 新记录置顶
            with open(history_file_path, "w", encoding="utf-8") as file:
                json.dump(history, file, indent=4)

    def load_history(self):
        # 加载历史文件记录
        history_file_path = self.get_history_file_path()
        if os.path.exists(history_file_path):
            with open(history_file_path, "r", encoding="utf-8") as file:
                return json.load(file)
        return []

    def create_menu(self):
        # 创建菜单栏
        menu_bar = self.menuBar()

        # 创建文件菜单
        file_menu = menu_bar.addMenu("文件")

        # 新建文件动作
        new_action = QAction("新建", self)
        new_action.triggered.connect(self.new_file)
        file_menu.addAction(new_action)

        # 打开文件动作
        open_action = QAction("打开", self)
        open_action.triggered.connect(self.load_excel)
        file_menu.addAction(open_action)

        # 保存文件动作
        save_action = QAction("保存", self)
        save_action.triggered.connect(self.save_excel)
        file_menu.addAction(save_action)

        # 创建历史文件菜单并保存引用
        self.history_menu = file_menu.addMenu("历史文件")
        self.update_history_menu()

    def update_history_menu(self):
        self.history_menu.clear()
        for file_path in self.history_files:
            history_action = QAction(file_path, self)
            history_action.triggered.connect(lambda checked, f=file_path: self.load_excel(f))
            self.history_menu.addAction(history_action)

    def load_history_file(self, file_path):
        # 根据历史路径打开文件
        self.load_excel(file_path)

    def new_file(self):
        # 清空当前表格数据
        self.table.setRowCount(10)
        self.table.setColumnCount(5)
        self.table.clearContents()

        # 更新状态栏
        self.statusBar.showMessage("已创建新文件")

        # 清空历史记录（如果需要）
        self.history_files.clear()

        # 更新历史文件菜单
        self.update_history_menu()

    from PyQt5.QtWidgets import QTableWidgetItem, QTableWidget, QInputDialog, QMenu, QAction
    from PyQt5.QtCore import Qt

    def contextMenuEvent(self, event):
        context_menu = QMenu(self)

        # 定义所有菜单项和动作
        actions = [
            ("复制", self.copy_cell),
            ("粘贴", self.paste_cell),
            ("删除选中项", self.clear_cell),
            ("批量新建行", self.new_rows),
            ("批量新建列", self.new_cols),
            ("修改列名", self.rename_column),
            ("批量填充", self.batch_fill),
            ("居中对齐", self.toggle_center_align),
            ("调整行间距和列间距", self.show_adjust_dialog)  # 添加弹窗触发菜单项
        ]

        # 添加菜单项并连接相应功能
        for action_text, action_func in actions:
            action = QAction(action_text, self)
            action.triggered.connect(action_func)
            context_menu.addAction(action)

        # 设置右键菜单的样式表，增加圆角效果
        context_menu.setStyleSheet("""
            QMenu {
                border: 1px solid #A0A0A0;
                border-radius: 5px;
                background-color: #f1f1f1;
                padding: 2px;
                margin: 2px;
            }
            QMenu::item {
                background-color: transparent;
                padding: 8px 20px;
                color: #000000;
                font-size: 14px;
                border-radius: 5px;
            }
            QMenu::item:selected {
                background-color: #0078d4;
                color: white;
                font-weight: bold;
            }
            QMenu::item:!selected {
                background-color: transparent;
            }
            QMenu::separator {
                height: 1px;
                background-color: #e1e1e1;
                margin: 4px 0;
            }
        """)

        # 显示菜单
        context_menu.exec_(event.globalPos())

    # 显示行列间距调整弹窗
    # 显示行列间距调整弹窗
    def show_adjust_dialog(self):
        dialog = RowColumnAdjustDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            row_start, row_end, col_start, col_end, row_height, col_width = dialog.get_values()
            if row_start is not None and col_start is not None:
                self.adjust_rows_and_columns(row_start, row_end, col_start, col_end, row_height, col_width)
                self.statusBar.showMessage("行列间距调整成功", 3000)

    # 批量调整行高和列宽
    def adjust_rows_and_columns(self, row_start, row_end, col_start, col_end, row_height, col_width):
        # 批量调整行高
        for row in range(row_start, row_end + 1):
            self.table.setRowHeight(row, row_height)

        # 批量调整列宽
        for col in range(col_start, col_end + 1):
            self.table.setColumnWidth(col, col_width)

    def toggle_center_align(self):
        # 获取选中的单元格范围
        selected_ranges = self.table.selectedRanges()

        if not selected_ranges:
            QMessageBox.warning(self, "警告", "没有选中任何单元格!")
            return

        # 遍历所有选中的范围
        for selection_range in selected_ranges:
            # 使用 QTableWidgetSelectionRange 的方法获取每个范围的边界
            for row in range(selection_range.topRow(), selection_range.bottomRow() + 1):
                for col in range(selection_range.leftColumn(), selection_range.rightColumn() + 1):
                    # 获取当前单元格
                    item = self.table.item(row, col)
                    if item:  # 确保该单元格有内容
                        current_alignment = item.textAlignment()

                        if current_alignment == Qt.AlignCenter:
                            # 如果已经是居中对齐，则设置为左对齐
                            item.setTextAlignment(Qt.AlignLeft)
                        else:
                            # 如果不是居中对齐，则设置为居中对齐
                            item.setTextAlignment(Qt.AlignCenter)

    def batch_fill(self):
        fill_value, ok = QInputDialog.getText(self, "批量填充", "输入填充的值:")
        if ok:

            # 获取选中的区域（ranges）
            selected_ranges = self.table.selectedRanges()

            # 输出选中的区域数

            if len(selected_ranges) == 0:
                return  # 如果没有选中任何区域，直接返回

            # 遍历所有选中的区域
            for irange in selected_ranges:
                # 获取区域的起始行列和结束行列
                start_row = irange.topRow()
                end_row = irange.bottomRow()
                start_col = irange.leftColumn()
                end_col = irange.rightColumn()

                # 遍历该区域的每个单元格并填充
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        item = self.table.item(row, col)
                        if item is None:
                            item = QTableWidgetItem()  # 如果单元格没有内容，创建新的QTableWidgetItem
                            self.table.setItem(row, col, item)
                        item.setText(fill_value)  # 设置文本值

            self.statusBar.showMessage(f"已批量填充选中的区域")

    def rename_column(self):
        # 获取选中的列索引
        selected_column = self.table.currentColumn()

        if selected_column != -1:
            # 获取当前列名
            current_column_name = self.table.horizontalHeaderItem(selected_column).text()

            # 弹出输入框，让用户输入新的列名
            new_column_name, ok = QInputDialog.getText(self, "修改列名", "请输入新的列名:", text=current_column_name)

            if ok and new_column_name:
                # 更新列名
                self.table.setHorizontalHeaderItem(selected_column, QTableWidgetItem(new_column_name))

                # 更新状态栏提示
                self.statusBar.showMessage(f"列名已修改为: {new_column_name}")



    def craft_row(self):
        self.insert_row()

    def delete_row(self):
        selected_ranges = self.table.selectedRanges()
        if selected_ranges:
            row = selected_ranges[0].topRow()
            self.table.removeRow(row)

    def insert_row(self):
        selected_ranges = self.table.selectedRanges()
        if selected_ranges:
            row = selected_ranges[0].topRow()
            self.table.insertRow(row)

    def copy_cell(self):
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            return

        copied_text = ""
        for row in range(selected_ranges[0].topRow(), selected_ranges[0].bottomRow() + 1):
            row_data = []
            for col in range(selected_ranges[0].leftColumn(), selected_ranges[0].rightColumn() + 1):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            copied_text += "\t".join(row_data) + "\n"

        QApplication.clipboard().setText(copied_text.strip())

    def paste_cell(self):
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            return

        paste_text = QApplication.clipboard().text()
        paste_rows = paste_text.splitlines()
        start_row = selected_ranges[0].topRow()
        start_col = selected_ranges[0].leftColumn()

        for row_offset, paste_row in enumerate(paste_rows):
            cells = paste_row.split("\t")
            for col_offset, cell_text in enumerate(cells):
                row = start_row + row_offset
                col = start_col + col_offset
                if row < self.table.rowCount() and col < self.table.columnCount():
                    item = self.table.item(row, col)
                    if not item:
                        item = QTableWidgetItem()
                        self.table.setItem(row, col, item)
                    item.setText(cell_text)

    def clear_cell(self):
        selected_items = self.table.selectedItems()
        if selected_items:
            # 遍历所有选中的单元格，将其内容设置为空
            for item in selected_items:
                item.setText("")  # 清空选中单元格的文本

    def load_excel(self, file_path=None):
        if not file_path:
            # 打开文件选择对话框
            file_path, _ = QFileDialog.getOpenFileName(self, "打开 Excel 文件", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            # 使用 openpyxl 打开 Excel 文件以读取样式
            try:
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active  # 获取活动工作表

                # 读取表格内容并转换为 DataFrame
                data = pd.DataFrame(sheet.values)

                # 确保列名正确
                data.columns = [f"列 {i + 1}" for i in range(data.shape[1])]  # 设置列名

                # 使用apply替换 NaT 值为空字符串
                data = data.applymap(lambda x: "" if pd.isna(x) else x)

                # 获取加粗和斜体字体单元格的位置
                bold_cells = []
                italic_cells = []
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.font.bold:
                            bold_cells.append((cell.row - 1, cell.column - 1))  # 行列索引从0开始
                        if cell.font.italic:
                            italic_cells.append((cell.row - 1, cell.column - 1))  # 行列索引从0开始

                # 更新表格并显示加粗和斜体字体
                self.display_data(data, bold_cells, italic_cells)

                # 更新状态栏提示
                self.statusBar.showMessage(f"文件已打开: {file_path}", 5000)  # 显示文件名，持续5秒
                self.setWindowTitle(f"简化版 Excel - {os.path.basename(file_path)}")

                # 更新历史文件
                if file_path not in self.history_files:
                    self.history_files.insert(0, file_path)  # 将新文件添加到历史列表
                    self.update_history_menu()

            except Exception as e:
                self.statusBar.showMessage(f"加载失败: {str(e)}", 5000)  # 显示加载失败提示

    def display_data(self, data, bold_cells=None, italic_cells=None):
        # 设置表格行列数
        self.table.setRowCount(data.shape[0])
        self.table.setColumnCount(data.shape[1])

        # 设置表头
        self.table.setHorizontalHeaderLabels(data.columns)

        # 填充数据
        for row in range(data.shape[0]):
            for col in range(data.shape[1]):
                item = QTableWidgetItem(str(data.iat[row, col]))

                # 设置字体样式：加粗
                if bold_cells and (row, col) in bold_cells:
                    font = item.font()
                    font.setBold(True)  # 设置加粗字体
                    item.setFont(font)

                # 设置字体样式：斜体
                if italic_cells and (row, col) in italic_cells:
                    font = item.font()
                    font.setItalic(True)  # 设置斜体字体
                    item.setFont(font)

                self.table.setItem(row, col, item)

        # 自适应行和列
        self.table.resizeColumnsToContents()  # 自适应列宽
        self.table.resizeRowsToContents()  # 自适应行高

    def save_excel(self):
        self.statusBar.showMessage("正在保存文件...")
        # 保存表格内容及格式到 Excel 文件
        path, _ = QFileDialog.getSaveFileName(self, "保存文件", "", "Excel 文件 (*.xlsx)")
        if not path:
            return

        # 创建新的 Excel 文件
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 遍历表格单元格，将内容和样式写入 Excel
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    cell = sheet.cell(row=row + 1, column=col + 1, value=item.text())

                    # 设置字体样式
                    font = Font(
                        bold=item.font().bold(),
                        italic=item.font().italic(),
                        underline="single" if item.font().underline() else None
                    )
                    cell.font = font

        # 保存到指定路径
        workbook.save(path)
        self.statusBar.showMessage(f"文件已保存: {path}", 5000)
    def path_save_excel(self, path):
        self.statusBar.showMessage("正在自动保存文件...")
        # 保存表格内容及格式到 Excel 文件
        if not path:
            return

        # 创建新的 Excel 文件
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 遍历表格单元格，将内容和样式写入 Excel
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    cell = sheet.cell(row=row + 1, column=col + 1, value=item.text())

                    # 设置字体样式
                    font = Font(
                        bold=item.font().bold(),
                        italic=item.font().italic(),
                        underline="single" if item.font().underline() else None
                    )
                    cell.font = font

        # 保存到指定路径
        workbook.save(path)
        self.statusBar.showMessage(f"文件已保存: {path}", 5000)

    def get_table_data(self):
        # 获取表格中的数据并转换为 pandas DataFrame
        rows = self.table.rowCount()
        cols = self.table.columnCount()
        data = {}

        # 获取表头
        headers = [self.table.horizontalHeaderItem(i).text() for i in range(cols)]

        for col in range(cols):
            col_data = []
            for row in range(rows):
                item = self.table.item(row, col)
                col_data.append(item.text() if item else "")
            data[headers[col]] = col_data

        # 转换为 DataFrame 并填充缺失值
        return pd.DataFrame(data).fillna("")


    def closeEvent(self, event):
        # 如果数据已被修改，则提示用户保存
        if self.data_modified:
            reply = QMessageBox.question(
                self,
                "保存更改",
                "是否保存更改？",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                self.save_excel()
                event.accept()  # 允许关闭
            elif reply == QMessageBox.No:
                event.accept()  # 允许关闭但不保存
            else:
                event.ignore()  # 取消关闭
        else:
            event.accept()  # 直接关闭

if __name__ == "__main__":
    app = QApplication(sys.argv)
    QLocale.setDefault(QLocale(QLocale.Chinese, QLocale.China))
    window = SimpleExcel()
    window.showMaximized()  # 启动时最大化，保留边框和标题栏
    window.show()
    sys.exit(app.exec_())

